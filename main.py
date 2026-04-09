
#!/usr/bin/env python3
"""
Daily Expense Tracker Telegram Bot
===================================
Features:
  - Add / Edit / Delete expenses
  - View summaries: Today, This Week, This Month, This Year
  - Export to Excel (.xlsx)

Requirements:
  pip install python-telegram-bot==20.7 openpyxl aiosqlite

Setup:
  1. Create a bot via @BotFather and copy the token.
  2. Set BOT_TOKEN below (or export as env var TELEGRAM_BOT_TOKEN).
  3. Run:  python expense_bot.py
"""

import os
import logging
import sqlite3
import io
from datetime import datetime, date, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ConversationHandler,
    ContextTypes,
    filters,
)

# ─── Config ───────────────────────────────────────────────────────────────────

BOT_TOKEN = os.getenv("8064044877:AAGdHp4ICm5Sk4XJbg5lkWcpaCRCNzw96X4", "8064044877:AAGdHp4ICm5Sk4XJbg5lkWcpaCRCNzw96X4")
DB_PATH = "expenses.db"

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ─── Conversation States ───────────────────────────────────────────────────────

(
    ADD_AMOUNT,
    ADD_CATEGORY,
    ADD_DESC,
    ADD_DATE,
    EDIT_CHOOSE,
    EDIT_FIELD,
    EDIT_VALUE,
    DELETE_CONFIRM,
) = range(8)

CATEGORIES = [
    "🍔 Food",
    "🚌 Transport",
    "🏠 Housing",
    "⚕️ Health",
    "🎉 Entertainment",
    "🛍️ Shopping",
    "📚 Education",
    "💡 Utilities",
    "💰 Savings",
    "📦 Other",
]

# ─── Database ──────────────────────────────────────────────────────────────────

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS expenses (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER NOT NULL,
            amount      REAL    NOT NULL,
            category    TEXT    NOT NULL,
            description TEXT,
            expense_date TEXT   NOT NULL,
            created_at  TEXT    NOT NULL
        )
    """)
    conn.commit()
    conn.close()


def db_add(user_id, amount, category, description, expense_date):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "INSERT INTO expenses (user_id, amount, category, description, expense_date, created_at) VALUES (?,?,?,?,?,?)",
        (user_id, amount, category, description, expense_date, datetime.now().isoformat()),
    )
    row_id = c.lastrowid
    conn.commit()
    conn.close()
    return row_id


def db_get(expense_id, user_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT * FROM expenses WHERE id=? AND user_id=?", (expense_id, user_id))
    row = c.fetchone()
    conn.close()
    return row


def db_update(expense_id, user_id, field, value):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(f"UPDATE expenses SET {field}=? WHERE id=? AND user_id=?", (value, expense_id, user_id))
    conn.commit()
    conn.close()


def db_delete(expense_id, user_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM expenses WHERE id=? AND user_id=?", (expense_id, user_id))
    conn.commit()
    conn.close()


def db_range(user_id, start: str, end: str):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "SELECT id, amount, category, description, expense_date FROM expenses "
        "WHERE user_id=? AND expense_date BETWEEN ? AND ? ORDER BY expense_date DESC",
        (user_id, start, end),
    )
    rows = c.fetchall()
    conn.close()
    return rows


def db_all(user_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "SELECT id, amount, category, description, expense_date FROM expenses "
        "WHERE user_id=? ORDER BY expense_date DESC",
        (user_id,),
    )
    rows = c.fetchall()
    conn.close()
    return rows


def db_recent(user_id, limit=10):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "SELECT id, amount, category, description, expense_date FROM expenses "
        "WHERE user_id=? ORDER BY expense_date DESC LIMIT ?",
        (user_id, limit),
    )
    rows = c.fetchall()
    conn.close()
    return rows

# ─── Helpers ───────────────────────────────────────────────────────────────────

def fmt_amount(amount: float) -> str:
    return f"${amount:,.2f}"


def today_str() -> str:
    return date.today().isoformat()


def week_range():
    today = date.today()
    start = today - timedelta(days=today.weekday())
    return start.isoformat(), today.isoformat()


def month_range():
    today = date.today()
    start = today.replace(day=1)
    return start.isoformat(), today.isoformat()


def year_range():
    today = date.today()
    start = today.replace(month=1, day=1)
    return start.isoformat(), today.isoformat()


def rows_summary(rows) -> str:
    if not rows:
        return "No expenses found."
    total = sum(r[1] for r in rows)
    by_cat: dict[str, float] = {}
    for r in rows:
        by_cat[r[2]] = by_cat.get(r[2], 0) + r[1]

    lines = [f"💳 *{len(rows)} expense(s) — Total: {fmt_amount(total)}*\n"]
    for cat, amt in sorted(by_cat.items(), key=lambda x: -x[1]):
        pct = amt / total * 100
        lines.append(f"  {cat}: {fmt_amount(amt)} ({pct:.0f}%)")
    lines.append(f"\n*Recent entries:*")
    for r in rows[:8]:
        lines.append(f"  `#{r[0]}` {r[4]}  {fmt_amount(r[1])}  {r[2]}  {r[3] or ''}")
    if len(rows) > 8:
        lines.append(f"  _...and {len(rows)-8} more_")
    return "\n".join(lines)


def main_keyboard():
    return ReplyKeyboardMarkup(
        [
            ["➕ Add Expense", "📋 View Today"],
            ["📅 This Week", "🗓️ This Month"],
            ["📆 This Year", "✏️ Edit Expense"],
            ["🗑️ Delete Expense", "📊 Export Excel"],
        ],
        resize_keyboard=True,
    )

# ─── Excel Export ──────────────────────────────────────────────────────────────

def build_excel(user_id: int, rows, title: str) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill    = PatternFill("solid", fgColor="D6E4F0")
    total_fill  = PatternFill("solid", fgColor="2E75B6")
    total_font  = Font(color="FFFFFF", bold=True, size=11)
    border      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    # Title row
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Expense Report — {title}"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # Generated
    ws.merge_cells("A2:F2")
    ws["A2"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="888888", size=10)
    ws["A2"].alignment = Alignment(horizontal="right")

    # Headers
    headers = ["#", "Date", "Category", "Description", "Amount"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[4].height = 20

    # Data
    total = 0.0
    by_cat: dict[str, float] = {}
    for i, r in enumerate(rows, 1):
        row_num = i + 4
        fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="EBF4FF")
        values = [r[0], r[4], r[2], r[3] or "", r[1]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = fill
            cell.border = border
            if col == 5:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = center if col <= 2 else Alignment(horizontal="left")
        total += r[1]
        by_cat[r[2]] = by_cat.get(r[2], 0) + r[1]

    # Total row
    total_row = len(rows) + 5
    ws.merge_cells(f"A{total_row}:D{total_row}")
    t = ws.cell(row=total_row, column=1, value="TOTAL")
    t.fill = total_fill; t.font = total_font; t.alignment = center; t.border = border
    amt = ws.cell(row=total_row, column=5, value=total)
    amt.fill = total_fill; amt.font = total_font
    amt.number_format = '"$"#,##0.00'; amt.alignment = Alignment(horizontal="right")
    amt.border = border

    # Category summary sheet
    ws2 = wb.create_sheet("Summary by Category")
    ws2["A1"].value = "Category"
    ws2["B1"].value = "Total Amount"
    ws2["C1"].value = "% Share"
    for col, hdr in enumerate(["Category", "Total Amount", "% Share"], 1):
        cell = ws2.cell(row=1, column=col, value=hdr)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = center; cell.border = border

    for i, (cat, amt_c) in enumerate(sorted(by_cat.items(), key=lambda x: -x[1]), 2):
        ws2.cell(row=i, column=1, value=cat).border = border
        ac = ws2.cell(row=i, column=2, value=amt_c)
        ac.number_format = '"$"#,##0.00'; ac.border = border
        pc = ws2.cell(row=i, column=3, value=round(amt_c / total * 100, 1) if total else 0)
        pc.number_format = "0.0\"%\""; pc.border = border

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 14
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─── Handlers ─────────────────────────────────────────────────────────────────

async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 *Welcome to your Expense Tracker Bot!*\n\n"
        "Track your daily spending right here in Telegram.\n\n"
        "Use the keyboard below or type /help.",
        parse_mode="Markdown",
        reply_markup=main_keyboard(),
    )


async def cmd_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = (
        "*Commands & Buttons*\n\n"
        "➕ *Add Expense* — record a new expense\n"
        "📋 *View Today* — today's expenses\n"
        "📅 *This Week* — current week\n"
        "🗓️ *This Month* — current month\n"
        "📆 *This Year* — current year\n"
        "✏️ *Edit Expense* — update any field by ID\n"
        "🗑️ *Delete Expense* — remove by ID\n"
        "📊 *Export Excel* — download full report\n\n"
        "You can also type /add, /today, /week, /month, /year, /export."
    )
    await update.message.reply_text(text, parse_mode="Markdown")


# ── Add flow ──────────────────────────────────────────────────────────────────

async def add_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    await update.message.reply_text(
        "💵 *Step 1/4 — Amount*\nEnter the expense amount (e.g. `12.50`):",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove(),
    )
    return ADD_AMOUNT


async def add_amount(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().replace("$", "").replace(",", "")
    try:
        amount = float(text)
        if amount <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❌ Invalid amount. Please enter a positive number:")
        return ADD_AMOUNT
    ctx.user_data["amount"] = amount
    keyboard = [[cat] for cat in CATEGORIES]
    await update.message.reply_text(
        "🏷️ *Step 2/4 — Category*\nChoose a category:",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True),
    )
    return ADD_CATEGORY


async def add_category(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    cat = update.message.text.strip()
    if cat not in CATEGORIES:
        await update.message.reply_text("Please choose a category from the list.")
        return ADD_CATEGORY
    ctx.user_data["category"] = cat
    await update.message.reply_text(
        "📝 *Step 3/4 — Description*\nAdd a short description (or type `skip`):",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove(),
    )
    return ADD_DESC


async def add_desc(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    ctx.user_data["description"] = "" if text.lower() == "skip" else text
    await update.message.reply_text(
        f"📅 *Step 4/4 — Date*\nEnter date as `YYYY-MM-DD` or type `today`:\n_(default: today = {today_str()})_",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup([["today"]], resize_keyboard=True, one_time_keyboard=True),
    )
    return ADD_DATE


async def add_date(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if text.lower() == "today":
        expense_date = today_str()
    else:
        try:
            datetime.strptime(text, "%Y-%m-%d")
            expense_date = text
        except ValueError:
            await update.message.reply_text("❌ Invalid format. Use YYYY-MM-DD or type `today`:")
            return ADD_DATE

    uid = update.effective_user.id
    row_id = db_add(
        uid,
        ctx.user_data["amount"],
        ctx.user_data["category"],
        ctx.user_data["description"],
        expense_date,
    )
    await update.message.reply_text(
        f"✅ *Expense saved!* (ID #{row_id})\n\n"
        f"  Amount:      {fmt_amount(ctx.user_data['amount'])}\n"
        f"  Category:    {ctx.user_data['category']}\n"
        f"  Description: {ctx.user_data['description'] or '—'}\n"
        f"  Date:        {expense_date}",
        parse_mode="Markdown",
        reply_markup=main_keyboard(),
    )
    ctx.user_data.clear()
    return ConversationHandler.END


# ── Edit flow ─────────────────────────────────────────────────────────────────

async def edit_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    rows = db_recent(uid, 10)
    if not rows:
        await update.message.reply_text("No expenses found.", reply_markup=main_keyboard())
        return ConversationHandler.END
    lines = ["✏️ *Edit Expense*\nRecent expenses:\n"]
    for r in rows:
        lines.append(f"`#{r[0]}` {r[4]}  {fmt_amount(r[1])}  {r[2]}  {r[3] or ''}")
    lines.append("\nEnter the *ID* of the expense to edit:")
    await update.message.reply_text(
        "\n".join(lines), parse_mode="Markdown", reply_markup=ReplyKeyboardRemove()
    )
    return EDIT_CHOOSE


async def edit_choose(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    try:
        eid = int(update.message.text.strip().lstrip("#"))
    except ValueError:
        await update.message.reply_text("Please enter a valid numeric ID:")
        return EDIT_CHOOSE
    uid = update.effective_user.id
    row = db_get(eid, uid)
    if not row:
        await update.message.reply_text("❌ Expense not found. Try again:")
        return EDIT_CHOOSE
    ctx.user_data["edit_id"] = eid
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("Amount", callback_data="ef_amount"),
         InlineKeyboardButton("Category", callback_data="ef_category")],
        [InlineKeyboardButton("Description", callback_data="ef_description"),
         InlineKeyboardButton("Date", callback_data="ef_expense_date")],
    ])
    await update.message.reply_text(
        f"Editing expense #{eid}:\n"
        f"  Amount: {fmt_amount(row[2])}  Category: {row[3]}\n"
        f"  Desc: {row[4] or '—'}  Date: {row[5]}\n\n"
        "Which field do you want to change?",
        reply_markup=kb,
    )
    return EDIT_FIELD


async def edit_field(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    field = query.data.replace("ef_", "")
    ctx.user_data["edit_field"] = field
    prompts = {
        "amount": "Enter new amount (e.g. `25.00`):",
        "category": "Enter new category:\n" + "\n".join(CATEGORIES),
        "description": "Enter new description (or `skip` to clear):",
        "expense_date": "Enter new date (YYYY-MM-DD):",
    }
    await query.edit_message_text(prompts[field], parse_mode="Markdown")
    return EDIT_VALUE


async def edit_value(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    eid = ctx.user_data["edit_id"]
    field = ctx.user_data["edit_field"]
    text = update.message.text.strip()

    if field == "amount":
        try:
            value = float(text.replace("$", "").replace(",", ""))
            if value <= 0:
                raise ValueError
        except ValueError:
            await update.message.reply_text("❌ Invalid amount. Try again:")
            return EDIT_VALUE
    elif field == "category":
        if text not in CATEGORIES:
            await update.message.reply_text("Please choose a valid category:")
            return EDIT_VALUE
        value = text
    elif field == "expense_date":
        try:
            datetime.strptime(text, "%Y-%m-%d")
            value = text
        except ValueError:
            await update.message.reply_text("❌ Use YYYY-MM-DD format:")
            return EDIT_VALUE
    else:
        value = "" if text.lower() == "skip" else text

    db_update(eid, uid, field, value)
    await update.message.reply_text(
        f"✅ Expense #{eid} updated ({field} → {value})",
        reply_markup=main_keyboard(),
    )
    ctx.user_data.clear()
    return ConversationHandler.END


# ── Delete flow ───────────────────────────────────────────────────────────────

async def delete_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    rows = db_recent(uid, 10)
    if not rows:
        await update.message.reply_text("No expenses found.", reply_markup=main_keyboard())
        return ConversationHandler.END
    lines = ["🗑️ *Delete Expense*\nRecent:\n"]
    for r in rows:
        lines.append(f"`#{r[0]}` {r[4]}  {fmt_amount(r[1])}  {r[2]}")
    lines.append("\nEnter the *ID* to delete:")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown", reply_markup=ReplyKeyboardRemove())
    return DELETE_CONFIRM


async def delete_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    try:
        eid = int(update.message.text.strip().lstrip("#"))
    except ValueError:
        await update.message.reply_text("Please enter a valid numeric ID:")
        return DELETE_CONFIRM
    uid = update.effective_user.id
    row = db_get(eid, uid)
    if not row:
        await update.message.reply_text("❌ Not found. Try again:", reply_markup=main_keyboard())
        return ConversationHandler.END
    db_delete(eid, uid)
    await update.message.reply_text(
        f"🗑️ Expense #{eid} deleted.", reply_markup=main_keyboard()
    )
    return ConversationHandler.END


# ── View handlers ─────────────────────────────────────────────────────────────

async def view_today(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    rows = db_range(uid, today_str(), today_str())
    await update.message.reply_text(
        f"📋 *Today's Expenses — {today_str()}*\n\n" + rows_summary(rows),
        parse_mode="Markdown",
        reply_markup=main_keyboard(),
    )


async def view_week(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s, e = week_range()
    rows = db_range(uid, s, e)
    await update.message.reply_text(
        f"📅 *This Week ({s} → {e})*\n\n" + rows_summary(rows),
        parse_mode="Markdown",
        reply_markup=main_keyboard(),
    )


async def view_month(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s, e = month_range()
    rows = db_range(uid, s, e)
    await update.message.reply_text(
        f"🗓️ *This Month ({s} → {e})*\n\n" + rows_summary(rows),
        parse_mode="Markdown",
        reply_markup=main_keyboard(),
    )


async def view_year(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s, e = year_range()
    rows = db_range(uid, s, e)
    await update.message.reply_text(
        f"📆 *This Year ({s} → {e})*\n\n" + rows_summary(rows),
        parse_mode="Markdown",
        reply_markup=main_keyboard(),
    )


# ── Export ────────────────────────────────────────────────────────────────────

async def export_excel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    rows = db_all(uid)
    if not rows:
        await update.message.reply_text("No expenses to export yet.", reply_markup=main_keyboard())
        return
    await update.message.reply_text("⏳ Generating your Excel report...")
    title = f"All time up to {today_str()}"
    buf = build_excel(uid, rows, title)
    filename = f"expenses_{today_str()}.xlsx"
    await update.message.reply_document(
        document=buf,
        filename=filename,
        caption=(
            f"📊 *Expense Report*\n"
            f"  {len(rows)} records exported\n"
            f"  Total: {fmt_amount(sum(r[1] for r in rows))}"
        ),
        parse_mode="Markdown",
        reply_markup=main_keyboard(),
    )


# ── Cancel ────────────────────────────────────────────────────────────────────

async def cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    await update.message.reply_text("Cancelled.", reply_markup=main_keyboard())
    return ConversationHandler.END


# ── Keyboard text router ───────────────────────────────────────────────────────

async def text_router(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "📋 View Today":
        await view_today(update, ctx)
    elif text == "📅 This Week":
        await view_week(update, ctx)
    elif text == "🗓️ This Month":
        await view_month(update, ctx)
    elif text == "📆 This Year":
        await view_year(update, ctx)
    elif text == "📊 Export Excel":
        await export_excel(update, ctx)
    else:
        await update.message.reply_text("Use the keyboard or /help.", reply_markup=main_keyboard())


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    init_db()
    app = Application.builder().token(BOT_TOKEN).build()

    add_conv = ConversationHandler(
        entry_points=[
            CommandHandler("add", add_start),
            MessageHandler(filters.Regex("^➕ Add Expense$"), add_start),
        ],
        states={
            ADD_AMOUNT:   [MessageHandler(filters.TEXT & ~filters.COMMAND, add_amount)],
            ADD_CATEGORY: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_category)],
            ADD_DESC:     [MessageHandler(filters.TEXT & ~filters.COMMAND, add_desc)],
            ADD_DATE:     [MessageHandler(filters.TEXT & ~filters.COMMAND, add_date)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    edit_conv = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex("^✏️ Edit Expense$"), edit_start),
        ],
        states={
            EDIT_CHOOSE: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_choose)],
            EDIT_FIELD:  [CallbackQueryHandler(edit_field, pattern="^ef_")],
            EDIT_VALUE:  [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_value)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    delete_conv = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex("^🗑️ Delete Expense$"), delete_start),
        ],
        states={
            DELETE_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, delete_confirm)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("today", view_today))
    app.add_handler(CommandHandler("week", view_week))
    app.add_handler(CommandHandler("month", view_month))
    app.add_handler(CommandHandler("year", view_year))
    app.add_handler(CommandHandler("export", export_excel))
    app.add_handler(add_conv)
    app.add_handler(edit_conv)
    app.add_handler(delete_conv)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_router))

    logger.info("Bot is running...")
    app.run_polling()


if __name__ == "__main__":
    main()